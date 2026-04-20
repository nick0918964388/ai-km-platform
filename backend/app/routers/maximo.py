"""
Maximo API Routes — NL→SQL structured query + knowledge base management.
"""

import logging

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
import io

_logger = logging.getLogger(__name__)

from app.db.session import get_db
from app.services.maximo_nl2sql import MaximoNL2SQL
from app.services.maximo_schema_rag import MaximoSchemaRAG
from app.services.maximo_doc_search import MaximoDocSearch
from app.auth import get_current_user, require_auth, require_admin

router = APIRouter(prefix="/maximo", tags=["Maximo"])

# ── Role normalizer ───────────────────────────────────────────────────────────

_VALID_ROLES = frozenset({"admin", "maint_manager", "maint_tech", "analyst", "viewer"})


def _normalize_role(raw_role: str | None) -> str:
    """Map unknown/guest/user roles to 'viewer'. UserContext accepts only known roles."""
    if raw_role in _VALID_ROLES:
        return raw_role
    return "viewer"


# ── NL→SQL ───────────────────────────────────────────────────────────────────

class NL2SQLRequest(BaseModel):
    question: str
    mode: str = "accurate"  # "fast" or "accurate"
    history: List[Dict[str, str]] = []  # [{"question": "...", "sql": "..."}, ...]


class NL2SQLResponse(BaseModel):
    success: bool
    sql: Optional[str] = None
    explanation: Optional[str] = None
    data: List[Any] = []
    columns: List[str] = []
    row_count: int = 0
    execution_ms: Optional[float] = None
    llm_ms: Optional[float] = None
    verify_ms: Optional[float] = None
    model: Optional[str] = None
    error: Optional[str] = None
    iterations: int = 1
    confidence: Optional[float] = None
    verification_history: List[Any] = []
    mode: str = "accurate"
    cached: bool = False
    query_plan: Optional[Any] = None
    chart_suggestion: Optional[Any] = None
    summary: Optional[str] = None
    clarification: Optional[Any] = None
    suggestions: List[Any] = []
    column_labels: Optional[Any] = None


@router.post("/schema/index")
async def index_schema(db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """索引 table catalog + attributes 到 Qdrant，供 RAG schema selector 使用。"""
    try:
        rag = MaximoSchemaRAG(db)
        stats = await rag.index_all()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"索引失敗：{str(e)}")


@router.post("/schema/discover")
async def discover_and_index(db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """自動偵測新的 maximo 表、產生 catalog entry、重新索引。"""
    try:
        rag = MaximoSchemaRAG(db)
        stats = await rag.discover_and_index()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"偵測/索引失敗：{str(e)}")


@router.post("/nl2sql")
async def nl2sql(req: NL2SQLRequest, db: AsyncSession = Depends(get_db), user: dict = Depends(require_auth)):
    """Convert natural language question to SQL and execute against Maximo tables.

    When MAXIMO_TOOL_ROUTER_ENABLED=true (default), the query first goes through
    MaximoQueryRouter. On a tool hit the response includes route_path/tool_name/rows.
    On fallback (or feature flag disabled), the response includes all original
    NL2SQLResponse fields for full backward compatibility.
    """
    from app.services.maximo_tools.feature_flag import is_tool_router_enabled

    # ── FAST PATH: feature flag disabled → pure legacy behaviour ──────────────
    if not is_tool_router_enabled():
        from app.services.intent_router import detect_ambiguity
        ambiguity = detect_ambiguity(req.question, history=req.history)
        if ambiguity:
            return NL2SQLResponse(
                success=False,
                explanation=ambiguity["message"],
                clarification=ambiguity,
            )
        service = MaximoNL2SQL(db)
        result = await service.query(req.question, mode=req.mode, user_context=user, conversation_history=req.history)
        return NL2SQLResponse(**result)

    # ── ROUTER PATH ───────────────────────────────────────────────────────────
    from app.services.maximo_tools.base import UserContext
    from app.services.maximo_tools.nl2sql_fallback import NL2SQLFallback
    from app.services.maximo_tools.router_factory import build_router
    from app.models.maximo_tool_schemas import serialize_response

    # query text: support both old "question" field and new "query" field
    query_text = req.question

    role = _normalize_role(user.get("role"))

    user_ctx = UserContext(
        user_id=str(user.get("id", "unknown")),
        role=role,
        section=user.get("section"),
        workshop=user.get("workshop"),
        email=user.get("email"),
    )

    _logger.info(
        "maximo tool call | user_id=%s role=%s section=%s workshop=%s query_len=%d",
        user_ctx.user_id,
        user_ctx.role,
        user_ctx.section,
        user_ctx.workshop,
        len(query_text),
    )

    # Zero-row policy: maint_tech without a section assignment must not receive
    # full-table results. This is a data-integrity problem — surface it clearly.
    if user_ctx.role == "maint_tech" and not user_ctx.section:
        _logger.warning(
            "maint_tech user %s has no section assignment — returning empty result set.",
            user_ctx.user_id,
        )
        return {
            "success": False,
            "data": [],
            "columns": [],
            "error": "您的帳號尚未設定段別（section），請聯絡管理員指派後再查詢。",
            "row_count": 0,
        }

    # NL2SQLFallback captures request-scoped db session
    fallback = NL2SQLFallback(db=db, user_context=user)
    query_router = build_router(fallback_fn=fallback)

    raw_result = await query_router.route(query_text, user_ctx)

    # ── Serialize new response (applies debug stripping by role) ──────────────
    response_data = serialize_response(raw_result, user_ctx.role)

    # ── Merge backward-compat fields ──────────────────────────────────────────
    # Old frontend clients expect: success, data, columns, sql, explanation, etc.
    # These come from different sources depending on route_path.
    route_path = raw_result.get("route_path")

    if route_path == "fallback" and fallback.last_result:
        # NL2SQL full result is available via side-channel
        nl2sql = fallback.last_result
        extra_compat = {
            "success": nl2sql.get("success", True),
            "sql": nl2sql.get("sql"),
            "explanation": nl2sql.get("explanation"),
            "data": nl2sql.get("data") or [],
            "columns": nl2sql.get("columns") or [],
            "execution_ms": nl2sql.get("execution_ms"),
            "llm_ms": nl2sql.get("llm_ms"),
            "verify_ms": nl2sql.get("verify_ms"),
            "model": nl2sql.get("model"),
            "error": nl2sql.get("error"),
            "iterations": nl2sql.get("iterations", 1),
            "confidence": nl2sql.get("confidence"),
            "verification_history": nl2sql.get("verification_history") or [],
            "mode": nl2sql.get("mode", req.mode),
            "cached": nl2sql.get("cached", False),
            "query_plan": nl2sql.get("query_plan"),
            "chart_suggestion": nl2sql.get("chart_suggestion"),
            "summary": nl2sql.get("summary"),
            "clarification": nl2sql.get("clarification"),
            "suggestions": nl2sql.get("suggestions") or [],
            "column_labels": nl2sql.get("column_labels"),
        }
    else:
        # tool / error paths: synthesise compat fields from rows
        rows = raw_result.get("rows", [])
        columns_list = list(rows[0].keys()) if rows else []
        extra_compat = {
            "success": route_path == "tool" or len(rows) > 0,
            "data": rows,
            "columns": columns_list,
            "sql": None,
            "explanation": None,
            "execution_ms": raw_result.get("elapsed_ms"),
            "llm_ms": None,
            "verify_ms": None,
            "model": None,
            "error": None,
            "iterations": 1,
            "confidence": None,
            "verification_history": [],
            "mode": req.mode,
            "cached": False,
            "query_plan": None,
            "chart_suggestion": raw_result.get("chart_hint"),
            "summary": None,
            "clarification": None,
            "suggestions": [],
            "column_labels": None,
        }

    # New fields override compat fields when both present
    merged = {**extra_compat, **response_data}
    return merged


@router.get("/export")
async def export_results(
    question: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Export NL→SQL query results as Excel (.xlsx)."""
    import openpyxl

    service = MaximoNL2SQL(db)
    result = await service.query(question, mode="fast", user_context=user)

    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "查詢失敗"))

    columns = result.get("columns", [])
    data = result.get("data", [])

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查詢結果"

    # Header row with styling
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F62FE", end_color="0F62FE", fill_type="solid")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = max(len(str(col_name)), max((len(str(row.get(col_name, ""))) for row in data[:100]), default=5))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"maximo_export.xlsx"
    encoded_filename = quote(f"maximo_export_{question[:20].replace(' ', '_')}.xlsx")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{encoded_filename}"},
    )


# ── Feedback ─────────────────────────────────────────────────────────────────

class FeedbackRequest(BaseModel):
    question: str
    sql: str
    rating: str  # "up" or "down"
    corrected_sql: Optional[str] = None


@router.post("/feedback")
async def submit_feedback(req: FeedbackRequest, db: AsyncSession = Depends(get_db), user: dict = Depends(require_auth)):
    """Submit feedback on NL→SQL result. 👍 auto-saves as verified example."""
    if req.rating == "up":
        # Save as verified few-shot example
        await db.execute(text("""
            INSERT INTO nl_sql_examples (question, sql_query, verified, tag)
            VALUES (:q, :sql, true, 'user_feedback')
            ON CONFLICT DO NOTHING
        """), {"q": req.question.strip(), "sql": req.sql.strip()})
        await db.commit()
        return {"success": True, "message": "已加入查詢範例庫"}

    elif req.rating == "down":
        if req.corrected_sql:
            # Validate corrected SQL is SELECT only
            if not req.corrected_sql.strip().lower().startswith("select"):
                raise HTTPException(status_code=400, detail="修正 SQL 必須是 SELECT 查詢")
            # Save corrected version
            await db.execute(text("""
                INSERT INTO nl_sql_examples (question, sql_query, verified, tag)
                VALUES (:q, :sql, true, 'user_corrected')
            """), {"q": req.question.strip(), "sql": req.corrected_sql.strip()})
            await db.commit()
            return {"success": True, "message": "已儲存修正的 SQL"}

        # Just record negative feedback (for analytics)
        await db.execute(text("""
            INSERT INTO query_audit_log (user_id, user_email, question, sql_generated, mode)
            VALUES (:uid, :email, :q, :sql, 'feedback_negative')
        """), {
            "uid": user.get("id", ""), "email": user.get("email", ""),
            "q": req.question, "sql": req.sql,
        })
        await db.commit()
        return {"success": True, "message": "已記錄回饋"}

    raise HTTPException(status_code=400, detail="rating 必須是 up 或 down")


# ── Knowledge Base Management ─────────────────────────────────────────────────

class RuleItem(BaseModel):
    id: int
    content: str
    tag: str = "general"

class ExampleItem(BaseModel):
    id: int
    question: str
    sql_query: str
    verified: bool
    tag: str = "general"

class KnowledgeResponse(BaseModel):
    rules: List[RuleItem]
    examples: List[ExampleItem]

class AddRuleRequest(BaseModel):
    content: str
    tag: str = "general"

class UpdateRuleRequest(BaseModel):
    content: str
    tag: str = "general"

class AddExampleRequest(BaseModel):
    question: str
    sql_query: str
    tag: str = "general"

class UpdateExampleRequest(BaseModel):
    question: str
    sql_query: str
    tag: str = "general"


@router.get("/knowledge", response_model=KnowledgeResponse)
async def get_knowledge(db: AsyncSession = Depends(get_db)):
    """Get all domain rules and SQL examples."""
    rules_result = await db.execute(text(
        "SELECT id, description, COALESCE(tag, 'general') as tag FROM maximo_field_metadata "
        "WHERE table_name = '_rules' ORDER BY tag, id"
    ))
    rules = [RuleItem(id=r.id, content=r.description, tag=r.tag) for r in rules_result.fetchall()]

    examples_result = await db.execute(text(
        "SELECT id, question, sql_query, verified, COALESCE(tag, 'general') as tag "
        "FROM nl_sql_examples ORDER BY tag, id"
    ))
    examples = [
        ExampleItem(id=r.id, question=r.question, sql_query=r.sql_query, verified=r.verified, tag=r.tag)
        for r in examples_result.fetchall()
    ]

    return KnowledgeResponse(rules=rules, examples=examples)


@router.post("/knowledge/rule", response_model=RuleItem)
async def add_rule(req: AddRuleRequest, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Add a domain knowledge rule."""
    if not req.content.strip():
        raise HTTPException(status_code=400, detail="Content cannot be empty")

    # Auto-generate unique column name using timestamp
    import time
    col_name = f"rule_{int(time.time() * 1000)}"

    result = await db.execute(text(
        "INSERT INTO maximo_field_metadata (table_name, column_name, display_name, description, tag) "
        "VALUES ('_rules', :col, '查詢規則', :desc, :tag) RETURNING id"
    ), {"col": col_name, "desc": req.content.strip(), "tag": req.tag})
    await db.commit()
    new_id = result.scalar()
    return RuleItem(id=new_id, content=req.content.strip(), tag=req.tag)


@router.delete("/knowledge/rule/{rule_id}")
async def delete_rule(rule_id: int, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Delete a domain rule."""
    result = await db.execute(text(
        "DELETE FROM maximo_field_metadata WHERE id = :id AND table_name = '_rules'"
    ), {"id": rule_id})
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"success": True}


@router.patch("/knowledge/rule/{rule_id}", response_model=RuleItem)
async def update_rule(rule_id: int, req: UpdateRuleRequest, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Update a domain rule."""
    if not req.content.strip():
        raise HTTPException(status_code=400, detail="Content cannot be empty")
    result = await db.execute(text(
        "UPDATE maximo_field_metadata SET description = :desc, tag = :tag "
        "WHERE id = :id AND table_name = '_rules' RETURNING id"
    ), {"desc": req.content.strip(), "tag": req.tag, "id": rule_id})
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return RuleItem(id=rule_id, content=req.content.strip(), tag=req.tag)


@router.post("/knowledge/example", response_model=ExampleItem)
async def add_example(req: AddExampleRequest, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Add a NL→SQL example."""
    if not req.question.strip() or not req.sql_query.strip():
        raise HTTPException(status_code=400, detail="Question and SQL cannot be empty")
    if not req.sql_query.strip().lower().startswith("select"):
        raise HTTPException(status_code=400, detail="Only SELECT queries allowed")

    result = await db.execute(text(
        "INSERT INTO nl_sql_examples (question, sql_query, verified, tag) "
        "VALUES (:q, :sql, true, :tag) RETURNING id"
    ), {"q": req.question.strip(), "sql": req.sql_query.strip(), "tag": req.tag})
    await db.commit()
    new_id = result.scalar()
    return ExampleItem(id=new_id, question=req.question.strip(), sql_query=req.sql_query.strip(), verified=True, tag=req.tag)


@router.delete("/knowledge/example/{example_id}")
async def delete_example(example_id: int, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Delete a SQL example."""
    result = await db.execute(text(
        "DELETE FROM nl_sql_examples WHERE id = :id"
    ), {"id": example_id})
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Example not found")
    return {"success": True}


@router.patch("/knowledge/example/{example_id}", response_model=ExampleItem)
async def update_example(example_id: int, req: UpdateExampleRequest, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Update a SQL example."""
    if not req.question.strip() or not req.sql_query.strip():
        raise HTTPException(status_code=400, detail="Question and SQL cannot be empty")
    if not req.sql_query.strip().lower().startswith("select"):
        raise HTTPException(status_code=400, detail="Only SELECT queries allowed")
    result = await db.execute(text(
        "UPDATE nl_sql_examples SET question = :q, sql_query = :sql, tag = :tag "
        "WHERE id = :id RETURNING id"
    ), {"q": req.question.strip(), "sql": req.sql_query.strip(), "tag": req.tag, "id": example_id})
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Example not found")
    return ExampleItem(id=example_id, question=req.question.strip(), sql_query=req.sql_query.strip(), verified=True, tag=req.tag)


# ── Related Documents (SOP) ──────────────────────────────────────────────────

class RelatedDocsRequest(BaseModel):
    wo_number: Optional[str] = None
    fault_code: Optional[str] = None
    description: Optional[str] = None
    asset_num: Optional[str] = None
    top_k: int = 5


@router.post("/related-docs")
async def find_related_docs(req: RelatedDocsRequest, db: AsyncSession = Depends(get_db), user: dict = Depends(require_auth)):
    """Find SOP documents related to a work order, fault code, or description."""
    service = MaximoDocSearch(db)
    result = await service.find_related_docs(
        wo_number=req.wo_number,
        fault_code=req.fault_code,
        description=req.description,
        asset_num=req.asset_num,
        top_k=req.top_k,
    )
    return result


class AddMappingRequest(BaseModel):
    fault_code: str
    document_id: str
    document_name: str
    fault_description: str = ""


@router.post("/related-docs/mapping")
async def add_doc_mapping(req: AddMappingRequest, db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Add a fault_code → document mapping (admin only)."""
    service = MaximoDocSearch(db)
    mapping_id = await service.add_mapping(
        fault_code=req.fault_code,
        document_id=req.document_id,
        document_name=req.document_name,
        fault_description=req.fault_description,
    )
    return {"success": True, "mapping_id": mapping_id}


# ── Audit ────────────────────────────────────────────────────────────────────

@router.get("/audit")
async def get_audit_log(db: AsyncSession = Depends(get_db), admin: dict = Depends(require_admin)):
    """Get recent NL→SQL query audit log. Requires admin."""
    result = await db.execute(text(
        "SELECT user_email, question, sql_generated, tables_accessed, row_count, created_at "
        "FROM query_audit_log ORDER BY created_at DESC LIMIT 50"
    ))
    logs = []
    for r in result.fetchall():
        logs.append({
            "user_email": r.user_email,
            "question": r.question,
            "sql": r.sql_generated,
            "tables": r.tables_accessed,
            "row_count": r.row_count,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"logs": logs}
